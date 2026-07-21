"""Fidelity run: GRASP-Parreno at parity of ITERATIONS, full comparison.

Runs our faithful GRASP re-implementation on the first N instances of
selected BR classes at a fixed iteration budget, then builds a table
comparing our mean fill against Parreno et al. (2008), Table 4 (first 10
instances, 1 run, 5000 iterations). The point is to show that AT PARITY
OF ITERATIONS the methods agree, so the wall-clock gap is throughput
(Python vs C++), not algorithm.

Design:
  * per-class iteration budget and instance count (heterogeneous classes
    cost far more per iteration, so fewer instances / a time cap there);
  * a per-instance wall-clock cap as a safety net so the run finishes;
  * parallel across instances (ProcessPoolExecutor), numba dominance
    kernel warmed once per worker;
  * CHECKPOINTED: results are written to JSON after every instance, so
    the run can be interrupted and resumed (just relaunch);
  * prints a comparison table at the end and writes results/fidelity.json.

Iteration completeness: each instance stops at max_iter OR at its
per-class time cap, whichever comes first. When the cap truncates a run,
that instance has fewer than max_iter iterations and is NOT comparable to
Parreno at the same iteration count. The summary table therefore reports,
per class, how many instances actually reached the target ("compl"
column) and computes the mean fill / gap ONLY on those. To force every
instance to the full iteration count (rigorous but unpredictable
duration), pass --guarantee-iters, which removes the time caps.

Usage (resume-safe; relaunch the same command to continue):
    python examples/fidelity_run.py --workers 8
    python examples/fidelity_run.py --workers 8 --guarantee-iters
    python examples/fidelity_run.py --workers 8 --iters 5000
    python examples/fidelity_run.py --summary results/fidelity.json

Defaults are sized so the whole run is a few hours on a modern i9.
Override per-class plan with --plan "thpack1:10:5000,thpack7:10:5000".
"""

import argparse
import json
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from gasp.io_utils import write_json_atomic  # noqa: E402

# Parreno et al. 2008, Table 4 (first 10 instances, 1 run, 5000 iter)
PARRENO_REF = {
    "thpack1": 92.95, "thpack2": 93.95, "thpack3": 93.54, "thpack4": 93.05,
    "thpack5": 93.01, "thpack6": 92.72, "thpack7": 91.62, "thpack8": 90.74,
    "thpack9": 90.43, "thpack10": 89.69, "thpack11": 89.29, "thpack12": 88.95,
    "thpack13": 88.22, "thpack14": 88.25, "thpack15": 88.23,
}

# default plan: (class, n_instances, max_iter, per-instance time cap s).
# Heterogeneous classes are heavier per iteration, so we cap time and
# rely on the convergence trend there rather than full 5000 iterations.
DEFAULT_PLAN = [
    ("thpack1", 10, 5000, 120),
    ("thpack7", 10, 5000, 600),
    ("thpack15", 5, 5000, 2400),
]


def _one(args):
    cls, idx, max_iter, cap, seed, solver = args
    from gasp.readers import load_set
    from gasp.ems_numba import warmup_dominance
    warmup_dominance()
    name, items, ks = load_set(cls, respect_orientation=True)[idx]
    if solver == "python":
        from gasp.grasp_parreno import solve_grasp_parreno
        r = solve_grasp_parreno(items, ks, max_iter=max_iter, seed=seed,
                                time_limit=cap)
    else:
        from gasp.java_backend import JavaGASP
        from gasp import GASPParams
        params = GASPParams(time_limit=cap or 0.0, max_iter=max_iter, seed=seed, allow_rotation=False)
        r = JavaGASP(items, ks, params, solver="gasp").run()
    return cls, name, round(r.best_fill, 2), r.iterations, round(r.elapsed, 1)


def parse_plan(text):
    plan = []
    for part in text.split(","):
        cls, n, it = part.split(":")
        cap = 3600
        plan.append((cls, int(n), int(it), cap))
    return plan


def summarize(res, target_iters=5000):
    print("\n" + "=" * 96)
    print(f"{'Classe':10s} {'n':>3} {'compl':>6} {'iter':>6} "
          f"{'nostro %':>9} {'Parreno %':>10} {'gap':>7} "
          f"{'t/ist(s)':>9} {'t/ist(min)':>11}")
    print("-" * 96)
    for cls in sorted(res, key=lambda c: int(c.replace("thpack", ""))):
        rows = res[cls]
        if not rows:
            continue
        # split complete (reached target_iters) from truncated runs; the
        # comparison with Parreno is only apples-to-apples on complete
        # runs, so the mean fill is computed on those alone.
        complete = [v for v in rows.values()
                    if v["iters"] >= target_iters]
        use = complete if complete else list(rows.values())
        fills = [v["fill"] for v in use]
        iters = [v["iters"] for v in use]
        times = [v.get("time", 0.0) for v in use]
        mean = sum(fills) / len(fills)
        ref = PARRENO_REF.get(cls)
        gap = (mean - ref) if ref else None
        avg_it = int(sum(iters) / len(iters))
        avg_t = sum(times) / len(times)
        gap_s = f"{gap:+.2f}" if gap is not None else "  -"
        ref_s = f"{ref:.2f}" if ref else "  -"
        # "compl" = how many of the n instances reached the target
        compl_s = f"{len(complete)}/{len(rows)}"
        flag = "" if complete else " *"
        print(f"{cls:10s} {len(rows):>3d} {compl_s:>6} {avg_it:>6d} "
              f"{mean:>8.2f}% {ref_s:>10} {gap_s:>7} "
              f"{avg_t:>8.1f} {avg_t/60:>10.2f}{flag}")
    print("=" * 96)
    print(f"Parreno: Tab.4, prime 10 ist., 1 run, {target_iters} iter.")
    print(f"'compl' = istanze che hanno raggiunto le {target_iters} "
          f"iterazioni; 'nostro %' e 'gap' sono calcolati SOLO su quelle")
    print("(confronto mela-a-mela). '*' = nessuna istanza completa: la")
    print("riga usa tutte le istanze ed e' indicativa, NON confrontabile.")
    print("'iter' = media iterazioni eseguite sulle istanze considerate.")
    print("'t/ist' al netto del warm-up numba; i tempi riflettono Python")
    print("vs il C++ di Parreno.")


def write_xlsx(res, path, target_iters=5000):
    """Write the fidelity comparison table to an Excel file: a Riepilogo
    sheet (one row per class, mean computed on complete instances only)
    and a Dettaglio sheet (one row per instance). Mirrors summarize()."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bold = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    normal = Font(name="Arial")
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"
    headers = ["Classe", "n", "Complete", "Iter medie", "Nostro fill %",
               "Parreno % (Tab.4)", "Gap (punti)", "t/ist (s)",
               "t/ist (min)", "Confrontabile"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
        c.fill = hdr_fill
        c.alignment = center
        c.border = border

    for cls in sorted(res, key=lambda c: int(c.replace("thpack", ""))):
        rows = res[cls]
        if not rows:
            continue
        complete = [v for v in rows.values()
                    if v["iters"] >= target_iters]
        use = complete if complete else list(rows.values())
        fills = [v["fill"] for v in use]
        iters = [v["iters"] for v in use]
        times = [v.get("time", 0.0) for v in use]
        mean = sum(fills) / len(fills)
        ref = PARRENO_REF.get(cls)
        gap = (mean - ref) if ref is not None else None
        comparable = bool(complete)
        row = [cls, len(rows), f"{len(complete)}/{len(rows)}",
               int(sum(iters)/len(iters)), round(mean, 2),
               (round(ref, 2) if ref is not None else "-"),
               (round(gap, 2) if gap is not None else "-"),
               round(sum(times)/len(times), 1),
               round(sum(times)/len(times)/60, 2),
               "si" if comparable else "NO (indicativo)"]
        ws.append(row)
        r = ws.max_row
        for c in ws[r]:
            c.font = normal
            c.border = border
            c.alignment = center
        if not comparable:
            for c in ws[r]:
                c.fill = warn_fill

    widths = [10, 5, 10, 11, 14, 17, 12, 10, 12, 16]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wd

    # note rows
    ws.append([])
    note = ws.max_row + 1
    ws.cell(row=note, column=1,
            value=(f"Media e gap calcolati SOLO sulle istanze che hanno "
                   f"raggiunto {target_iters} iterazioni (confronto "
                   f"mela-a-mela). Righe gialle: nessuna istanza completa, "
                   f"valore indicativo. Parreno: Tab.4, 1 run, "
                   f"{target_iters} iter. Tempi al netto del warm-up "
                   f"numba; riflettono Python vs C++."))
    ws.cell(row=note, column=1).font = Font(name="Arial", italic=True,
                                            size=9)

    # detail sheet
    ws2 = wb.create_sheet("Dettaglio")
    dh = ["Classe", "Istanza", "Iterazioni", "Completa", "Fill %",
          "Tempo (s)"]
    ws2.append(dh)
    for c in ws2[1]:
        c.font = bold
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    for cls in sorted(res, key=lambda c: int(c.replace("thpack", ""))):
        for name, v in sorted(res[cls].items()):
            done = v["iters"] >= target_iters
            ws2.append([cls, name, v["iters"], "si" if done else "no",
                        round(v["fill"], 2), round(v.get("time", 0.0), 1)])
            for c in ws2[ws2.max_row]:
                c.font = normal
                c.border = border
                c.alignment = center
            if not done:
                for c in ws2[ws2.max_row]:
                    c.fill = warn_fill
    for i, wd in enumerate([10, 16, 11, 10, 9, 10], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = wd

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--iters", type=int, default=None,
                    help="override iteration budget for every class")
    ap.add_argument("--guarantee-iters", action="store_true",
                    help="rimuove il cap di tempo: ogni istanza arriva "
                         "DAVVERO al numero di iterazioni richiesto "
                         "(confronto rigoroso, ma durata imprevedibile)")
    ap.add_argument("--plan", type=str, default=None,
                    help='e.g. "thpack1:10:5000,thpack7:10:5000"')
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--out", type=str, default="results/fidelity.json")
    ap.add_argument("--summary", type=str, default=None)
    ap.add_argument("--solver", choices=["java", "python"], default="java",
                    help="solver backend (default: java)")
    ap.add_argument("--xlsx", type=str, default=None,
                    help="scrive la tabella di fedelta' anche in un file "
                         "Excel al percorso indicato (es. "
                         "results/fidelity.xlsx)")
    args = ap.parse_args()

    plan = parse_plan(args.plan) if args.plan else list(DEFAULT_PLAN)
    if args.iters:
        plan = [(c, n, args.iters, cap) for (c, n, _it, cap) in plan]
    # the iteration target (for the summary's completeness check) is the
    # max_iter common to the plan; assumes a single target across classes
    target_iters = plan[0][2] if plan else 5000
    if args.guarantee_iters:
        # drop the time caps so every instance truly reaches max_iter
        plan = [(c, n, it, None) for (c, n, it, _cap) in plan]

    if args.summary:
        data = json.loads(Path(args.summary).read_text())
        summarize(data, target_iters)
        if args.xlsx:
            p = write_xlsx(data, args.xlsx, target_iters)
            print(f"\nTabella Excel scritta: {p}")
        return

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    res = json.loads(out.read_text()) if out.exists() else {}

    # build task list, skipping already-done instances (resume)
    from gasp.readers import load_set
    tasks = []
    for (cls, n, max_iter, cap) in plan:
        res.setdefault(cls, {})
        names = [nm for (nm, _i, _k) in
                 load_set(cls, respect_orientation=True)[:n]]
        for idx, nm in enumerate(names):
            if nm in res[cls]:
                continue
            tasks.append((cls, idx, max_iter, cap, args.seed, args.solver))

    print(f"Fidelity run: {len(tasks)} istanze da eseguire "
          f"({sum(len(v) for v in res.values())} gia' in checkpoint).")
    if not tasks:
        print("Tutto gia' calcolato. Riepilogo:")
        summarize(res, target_iters)
        if args.xlsx:
            print(f"\nTabella Excel scritta: "
                  f"{write_xlsx(res, args.xlsx, target_iters)}")
        return

    t0 = time.time()
    if args.workers <= 1:
        for k, t in enumerate(tasks):
            cls, name, fill, iters, el = _one(t)
            res[cls][name] = {"fill": fill, "iters": iters, "time": el}
            write_json_atomic(out, res)
            print(f"[{k+1}/{len(tasks)}] {cls}/{name}: {fill}% "
                  f"({iters} iter, {el}s)", flush=True)
    else:
        from concurrent.futures import ProcessPoolExecutor, as_completed
        print(f"Esecuzione PARALLELA su {min(args.workers, len(tasks))} "
              f"worker (i completamenti compaiono a ondate).", flush=True)
        with ProcessPoolExecutor(max_workers=args.workers) as ex:
            futs = {ex.submit(_one, t): t for t in tasks}
            for k, fut in enumerate(as_completed(futs)):
                cls, name, fill, iters, el = fut.result()
                res[cls][name] = {"fill": fill, "iters": iters, "time": el}
                write_json_atomic(out, res)
                print(f"[{k+1}/{len(tasks)}] {cls}/{name}: {fill}% "
                      f"({iters} iter, {el}s)", flush=True)

    print(f"\nCompletato in {(time.time()-t0)/60:.1f} min. "
          f"Checkpoint: {out}")
    summarize(res, target_iters)
    if args.xlsx:
        print(f"\nTabella Excel scritta: "
              f"{write_xlsx(res, args.xlsx, target_iters)}")


if __name__ == "__main__":
    import multiprocessing as _mp
    _mp.freeze_support()
    main()
