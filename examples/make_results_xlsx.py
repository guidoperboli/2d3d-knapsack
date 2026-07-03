#!/usr/bin/env python3
"""Build the Excel results workbook from run_campaign.py checkpoints.

Reads one or more campaign JSON files (the checkpoint format written by
examples/run_campaign.py), merges them, and produces a formatted .xlsx:
one sheet per set with per-instance rows, per-seed runs, best/mean and
gap or fill computed as live Excel formulas (recalculated by Excel or
LibreOffice on open), proven optima highlighted in green, and a summary
sheet with cross-sheet formulas.

Examples
--------
python examples/make_results_xlsx.py
python examples/make_results_xlsx.py --input results/campaign.json \
    results/br_bestof10.json --out results/Risultati_GASP.xlsx

Requires openpyxl (pip install openpyxl).
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10)
FILL = PatternFill("solid", start_color="2E5395")
F = Font(name="Arial", size=10)
FB = Font(name="Arial", size=10, bold=True)
GREEN = PatternFill("solid", start_color="C6EFCE")


def merge_inputs(paths) -> dict:
    res: dict = {}
    for p in paths:
        data = json.loads(Path(p).read_text())
        # results produced by the ALNS solver live in files named
        # *alns*.json; suffix their set names so they get their own
        # sheets ("thpack7 (ALNS)") instead of overwriting the GASP ones.
        is_alns = "alns" in Path(p).name.lower()
        for set_name, insts in data.items():
            key = f"{set_name} (ALNS)" if is_alns else set_name
            dst = res.setdefault(key, {})
            for name, rec in insts.items():
                if name in dst:
                    dst[name]["runs"].update(rec.get("runs", {}))
                else:
                    dst[name] = rec
    return res


def head(ws, headers, widths):
    ws.append(headers)
    for c, w in enumerate(widths, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR
        cell.fill = FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w


def build(res: dict, out: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Riepilogo"
    head(summary, ["Set", "Istanze", "Seed", "Metrica",
                   "Mean (medio)", "Std (medio)", "Nota", "Ottimi"],
         [12, 9, 6, 22, 14, 12, 24, 8])

    for set_name in sorted(res):
        insts = res[set_name]
        n_seeds = max((len(r.get("runs", {})) for r in insts.values()),
                      default=0)
        if n_seeds == 0:
            continue
        metric = next(iter(insts.values()))["metric"]
        is_fill = metric == "fill"

        ws = wb.create_sheet(set_name[:31])
        seed_cols = [f"S{k}" for k in range(1, n_seeds + 1)]
        if is_fill:
            headers = (["Istanza", "n"] + seed_cols
                       + ["Mean fill %", "Std", "Min %", "Max %",
                          "Best-of (NO conf.)", "Tempo medio (s)"])
            widths = [16, 6] + [9] * n_seeds + [12, 7, 9, 9, 17, 14]
        else:
            headers = (["Istanza", "n", "Rif. (" + metric + ")"]
                       + seed_cols
                       + ["Best profit", "Gap best %", "Gap mean %",
                          "Gap S1 %", "Tempo medio (s)"])
            widths = [16, 6, 13] + [11] * n_seeds + [12, 11, 11, 10, 14]
        head(ws, headers, widths)

        r = 1
        for name in sorted(insts):
            rec = insts[name]
            runs = rec.get("runs", {})
            if not runs:
                continue
            r += 1
            vals = [runs.get(str(k)) for k in range(1, n_seeds + 1)]
            tr = list(rec.get("time_runs", {}).values())
            tmean = round(sum(tr) / len(tr), 3) if tr else None
            first = get_column_letter(3 if is_fill else 4)
            last = get_column_letter((2 if is_fill else 3) + n_seeds)
            rng = f"{first}{r}:{last}{r}"
            if is_fill:
                # Mean is the honest comparison statistic; best-of is
                # flagged as not-for-comparison (cherry-pick).
                ws.append([name, rec.get("n", "")] + vals
                          + [f"=ROUND(AVERAGE({rng}),2)",
                             f"=ROUND(STDEVP({rng}),2)",
                             f"=ROUND(MIN({rng}),2)",
                             f"=ROUND(MAX({rng}),2)",
                             f"=ROUND(MAX({rng}),2)", tmean])
            else:
                ref = rec["ref"]
                bcol = get_column_letter(3 + n_seeds + 1)
                ws.append([name, rec.get("n", ""), ref] + vals
                          + [f"=MAX({rng})",
                             f"=ROUND(100*($C{r}-{bcol}{r})/$C{r},2)",
                             f"=ROUND(100*($C{r}-AVERAGE({rng}))/$C{r},2)",
                             f'=IF(D{r}="","",ROUND(100*($C{r}-D{r})/$C{r},2))',
                             tmean])
                if metric == "opt" and max(v for v in vals if v is not None) >= ref:
                    for c in ws[r]:
                        c.fill = GREEN
            for c in ws[r]:
                c.font = F

        # riga media del foglio
        r += 1
        if is_fill:
            mc = get_column_letter(2 + n_seeds + 1)   # Mean
            sc = get_column_letter(2 + n_seeds + 2)   # Std
            mn = get_column_letter(2 + n_seeds + 3)   # Min
            mx = get_column_letter(2 + n_seeds + 4)   # Max
            bo = get_column_letter(2 + n_seeds + 5)   # Best-of
            tc = get_column_letter(2 + n_seeds + 6)   # Tempo
            ws.append(["MEDIA", ""] + [""] * n_seeds
                      + [f"=ROUND(AVERAGE({mc}2:{mc}{r-1}),2)",
                         f"=ROUND(AVERAGE({sc}2:{sc}{r-1}),2)",
                         f"=ROUND(AVERAGE({mn}2:{mn}{r-1}),2)",
                         f"=ROUND(AVERAGE({mx}2:{mx}{r-1}),2)",
                         f"=ROUND(AVERAGE({bo}2:{bo}{r-1}),2)",
                         f"=ROUND(AVERAGE({tc}2:{tc}{r-1}),3)"])
        else:
            g1 = get_column_letter(3 + n_seeds + 2)
            g2 = get_column_letter(3 + n_seeds + 3)
            g3 = get_column_letter(3 + n_seeds + 4)
            tc = get_column_letter(3 + n_seeds + 5)   # Tempo
            ws.append(["MEDIA", "", ""] + [""] * n_seeds
                      + ["", f"=ROUND(AVERAGE({g1}2:{g1}{r-1}),2)",
                         f"=ROUND(AVERAGE({g2}2:{g2}{r-1}),2)",
                         f"=ROUND(AVERAGE({g3}2:{g3}{r-1}),2)",
                         f"=ROUND(AVERAGE({tc}2:{tc}{r-1}),3)"])
        for c in ws[r]:
            c.font = FB

        # riga nel riepilogo (formule cross-foglio)
        sheet_ref = f"'{ws.title}'"
        if is_fill:
            mc = get_column_letter(2 + n_seeds + 1)   # Mean
            sc = get_column_letter(2 + n_seeds + 2)    # Std
            # last data row is r-1 (the MEDIA row is r), values per
            # instance are rows 2..r-1
            summary.append([
                set_name, r - 2, n_seeds, "fill % (mean su seed)",
                f"=ROUND(AVERAGE({sheet_ref}!{mc}2:{mc}{r-1}),2)",
                f"=ROUND(AVERAGE({sheet_ref}!{sc}2:{sc}{r-1}),2)",
                "mean/std sui seed fissi", ""])
        else:
            g1 = get_column_letter(3 + n_seeds + 2)
            g2 = get_column_letter(3 + n_seeds + 3)
            n_opt = sum(
                1 for rec in insts.values()
                if rec["metric"] == "opt" and rec.get("runs")
                and max(rec["runs"].values()) >= rec["ref"])
            label = ("gap % vs ottimo" if metric == "opt"
                     else "gap % vs UB 1D")
            g3 = get_column_letter(3 + n_seeds + 4)
            summary.append([
                set_name, r - 2, n_seeds, label,
                f"=ROUND(AVERAGE({sheet_ref}!{g3}2:{g3}{r-1}),2)",
                f"=ROUND(AVERAGE({sheet_ref}!{g1}2:{g1}{r-1}),2)",
                f"=ROUND(AVERAGE({sheet_ref}!{g2}2:{g2}{r-1}),2)",
                n_opt if metric == "opt" else ""])
        for c in summary[summary.max_row]:
            c.font = F
        summary.cell(row=summary.max_row, column=1).font = FB

    # ---- stage-decomposition sheet (3D fill sets with seed/pre-layout)
    import statistics as _st
    decomp_rows = []
    for set_name in sorted(res):
        for name, rec in res[set_name].items():
            if rec.get("metric") != "fill":
                continue
            sruns = rec.get("seed_runs")
            pruns = rec.get("prelayout_runs")
            fruns = rec.get("runs")
            if not sruns or not fruns:
                continue
            seeds = sorted(sruns.keys(), key=lambda s: int(s))
            sv = [sruns[s] for s in seeds if sruns.get(s) is not None]
            pv = [pruns.get(s) for s in seeds if pruns.get(s) is not None]
            fv = [fruns.get(s) for s in seeds if fruns.get(s) is not None]
            if not sv or not fv:
                continue
            seed_m = _st.mean(sv)
            pre_m = _st.mean(pv) if pv else seed_m
            fin_m = _st.mean(fv)
            decomp_rows.append((set_name, name, seed_m, pre_m, fin_m))

    if decomp_rows:
        ds = wb.create_sheet("Scomposizione")
        head(ds, ["Set", "Istanza", "Seme %", "+GASP %", "Finale %",
                  "d(GASP)", "d(layout)"],
             [10, 16, 9, 9, 9, 9, 9])
        r = 1
        for (sett, name, sm, pm, fm) in decomp_rows:
            r += 1
            ds.append([sett, name,
                       round(sm, 2), round(pm, 2), round(fm, 2),
                       f"=ROUND(D{r}-C{r},2)", f"=ROUND(E{r}-D{r},2)"])
            for c in ds[r]:
                c.font = F
        # media per set + media globale
        r += 1
        ds.append(["MEDIA", "",
                   f"=ROUND(AVERAGE(C2:C{r-1}),2)",
                   f"=ROUND(AVERAGE(D2:D{r-1}),2)",
                   f"=ROUND(AVERAGE(E2:E{r-1}),2)",
                   f"=ROUND(AVERAGE(F2:F{r-1}),2)",
                   f"=ROUND(AVERAGE(G2:G{r-1}),2)"])
        for c in ds[r]:
            c.font = FB
        ds.append([])
        ds.append(["Seme = costruttivo Parreno; +GASP = dopo il loop "
                   "con learning; Finale = dopo la ricerca locale a 5 "
                   "vicinati. Medie sui seed fissati a priori."])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Scritto: {out}")
    print("Le formule vengono ricalcolate da Excel/LibreOffice "
          "all'apertura del file.")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", nargs="+",
                    default=["results/campaign.json"],
                    help="uno o piu' file JSON di campagna (vengono fusi)")
    ap.add_argument("--out", default="results/Risultati_GASP.xlsx",
                    help="file Excel di output")
    args = ap.parse_args()
    res = merge_inputs(args.input)
    if not res:
        raise SystemExit("Nessun risultato trovato nei file di input.")
    build(res, Path(args.out))


if __name__ == "__main__":
    main()
